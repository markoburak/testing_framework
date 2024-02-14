import openpyxl
def get_data_from_excel(path: str, sheet_name: str) -> list:

    final_data = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    total_rows = get_row_count(path, sheet_name)
    total_columns = get_column_count(path, sheet_name)

    for row in range(2, total_rows + 1):
        row_data = []
        for column in range(1, total_columns + 1):
            row_data.append(sheet.cell(row=row, column=column).value)
        final_data.append(row_data)

    return final_data

def get_row_count(path: str, sheet_name: str) -> int:
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name]
    return ws.max_row


def get_column_count(path: str, sheet_name: str) -> int:
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name]
    return ws.max_column


def get_cell_value(path: str, sheet_name: str, row: int, column: int) -> str:
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name]
    return ws.cell(row=row, column=column).value


def set_cell_value(path: str, sheet_name: str, row: int, column: int, value: str) -> None:
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name]
    ws.cell(row=row, column=column).value = value
    wb.save(path)
